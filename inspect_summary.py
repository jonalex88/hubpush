import openpyxl
p = r'c:\Vibes\Hubpush\output v2 all fields FULL.summary.xlsx'
wb = openpyxl.load_workbook(p, data_only=True)
for ws in wb.worksheets:
    print(ws.title, ws.max_row, ws.max_column)
    for r in ws.iter_rows(min_row=1, max_row=min(ws.max_row,5), values_only=True):
        print(r)
    print('---')
wb.close()

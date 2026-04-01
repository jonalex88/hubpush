"""
HubSpot push test - processes first PASS/PASS row from output workbook.

Steps:
  1. Load first PASS/PASS row from output workbook
  2. Resolve mandate + bank proof files from subfolder
  3. Create or find Contact by email (dedup)
  4. Create or find Company by registration number (dedup)
  5. Associate Contact -> Company
  6. Upload mandate file to HubSpot Files, get URL
  7. Upload bank proof file to HubSpot Files, get URL
  8. Create Deal with all properties + file URLs
  9. Associate Deal -> Company

Run:  .\.venv\Scripts\python.exe hs_push_test.py
"""
import json
import mimetypes
import os
import sys
import urllib.request
import urllib.error
from pathlib import Path

import openpyxl

TOKEN = os.getenv("HUBSPOT_API_TOKEN", "")
BASE = "https://api.hubapi.com"
ROOT = Path(r"c:\Vibes\Hubpush\1.FBEO\1.FBEO")
OUTPUT = Path(r"c:\Vibes\Hubpush\output v2 all fields FULL.resume.xlsx")
PORTAL_ID = "145268660"

AUTH_HEADERS = {
    "Authorization": f"Bearer {TOKEN}",
    "Content-Type": "application/json",
}

MANDATE_KEYS = ("mandate", "debit", "order")
BANK_KEYS = ("bank", "account", "confirmation", "proof", "statement")
SUPPORTED_DOC_EXTS = {".pdf", ".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp"}

DEAL_PIPELINE = "default"
DEAL_STAGE = "appointmentscheduled"
DEAL_OWNER_ID = "29362124"

# Standard HubSpot v4 association type IDs (HUBSPOT_DEFINED)
ASSOC_CONTACT_TO_COMPANY = 279   # contact primary company
ASSOC_DEAL_TO_COMPANY = 341      # deal primary company


# ── HTTP helpers ─────────────────────────────────────────────────────────────

def _log_request(method: str, path: str) -> None:
    print(f"   --> {method} {BASE}{path}")


def api_get(path: str) -> dict:
    _log_request("GET", path)
    req = urllib.request.Request(f"{BASE}{path}", headers=AUTH_HEADERS)
    try:
        with urllib.request.urlopen(req) as resp:
            return json.loads(resp.read())
    except urllib.error.HTTPError as e:
        print(f"       HTTP {e.code}: {e.read().decode()[:600]}")
        return {}


def api_post(path: str, data: dict) -> dict:
    _log_request("POST", path)
    body = json.dumps(data).encode()
    req = urllib.request.Request(f"{BASE}{path}", data=body, headers=AUTH_HEADERS, method="POST")
    try:
        with urllib.request.urlopen(req) as resp:
            return json.loads(resp.read())
    except urllib.error.HTTPError as e:
        print(f"       HTTP {e.code}: {e.read().decode()[:600]}")
        return {}


def api_put(path: str, data: list | dict | None = None) -> bool:
    _log_request("PUT", path)
    body = json.dumps(data).encode() if data is not None else b"[]"
    req = urllib.request.Request(f"{BASE}{path}", data=body, headers=AUTH_HEADERS, method="PUT")
    try:
        with urllib.request.urlopen(req) as _:
            return True
    except urllib.error.HTTPError as e:
        print(f"       HTTP {e.code}: {e.read().decode()[:600]}")
        return False


def upload_file(file_path: Path, folder_path: str = "/hubpush") -> str | None:
    """
    Upload a file to HubSpot Files v3 API.
    Returns the public URL of the uploaded file, or None on failure.
    """
    print(f"   --> POST {BASE}/files/v3/files  [{file_path.name}]")
    mime_type = mimetypes.guess_type(str(file_path))[0] or "application/octet-stream"
    file_bytes = file_path.read_bytes()

    boundary = "HubPushBoundary7MA4YWxkTrZu0gW"
    CRLF = b"\r\n"

    def part_text(name: str, value: str) -> bytes:
        return (
            f"--{boundary}\r\n"
            f'Content-Disposition: form-data; name="{name}"\r\n\r\n'
            f"{value}"
        ).encode() + CRLF

    options = json.dumps({
        "access": "PUBLIC_INDEXABLE",
        "overwrite": False,
        "duplicateValidationStrategy": "NONE",
        "duplicateValidationScope": "ENTIRE_PORTAL",
    })

    body = (
        part_text("folderPath", folder_path)
        + part_text("options", options)
        + (
            f"--{boundary}\r\n"
            f'Content-Disposition: form-data; name="file"; filename="{file_path.name}"\r\n'
            f"Content-Type: {mime_type}\r\n\r\n"
        ).encode()
        + file_bytes
        + CRLF
        + f"--{boundary}--\r\n".encode()
    )

    upload_headers = {
        "Authorization": f"Bearer {TOKEN}",
        "Content-Type": f"multipart/form-data; boundary={boundary}",
    }

    req = urllib.request.Request(
        f"{BASE}/files/v3/files",
        data=body,
        headers=upload_headers,
        method="POST",
    )
    try:
        with urllib.request.urlopen(req) as resp:
            result = json.loads(resp.read())
            file_url = result.get("url") or result.get("id")
            print(f"       Uploaded OK -> {file_url}")
            return str(file_url) if file_url else None
    except urllib.error.HTTPError as e:
        print(f"       HTTP {e.code}: {e.read().decode()[:600]}")
        return None


# ── HubSpot object helpers ────────────────────────────────────────────────────

def find_contact_by_email(email: str) -> str | None:
    resp = api_post("/crm/v3/objects/contacts/search", {
        "filterGroups": [{"filters": [
            {"propertyName": "email", "operator": "EQ", "value": email}
        ]}],
        "properties": ["email", "firstname"],
        "limit": 1,
    })
    results = resp.get("results", [])
    if results:
        print(f"       Found existing contact: id={results[0]['id']}")
        return results[0]["id"]
    return None


def find_company_by_registration(reg_no: str) -> str | None:
    resp = api_post("/crm/v3/objects/companies/search", {
        "filterGroups": [{"filters": [
            {"propertyName": "company_registration_no_", "operator": "EQ", "value": reg_no}
        ]}],
        "properties": ["name", "company_registration_no_"],
        "limit": 1,
    })
    results = resp.get("results", [])
    if results:
        print(f"       Found existing company: id={results[0]['id']}")
        return results[0]["id"]
    return None


def associate_objects(
    from_type: str, from_id: str,
    to_type: str, to_id: str,
    assoc_type_id: int,
) -> bool:
    path = f"/crm/v4/objects/{from_type}/{from_id}/associations/{to_type}/{to_id}"
    payload = [{"associationCategory": "HUBSPOT_DEFINED", "associationTypeId": assoc_type_id}]
    return api_put(path, payload)


# ── Data helpers ──────────────────────────────────────────────────────────────

def get_first_pass_row() -> dict:
    wb = openpyxl.load_workbook(str(OUTPUT), data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    for r in range(2, ws.max_row + 1):
        row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        if (
            row.get("debit order mandate validation result") == "PASS"
            and row.get("bank account proof validation") == "PASS"
        ):
            wb.close()
            return row
    wb.close()
    raise RuntimeError("No PASS/PASS row found in output workbook.")


def classify_folder_docs(folder: Path) -> tuple[Path | None, Path | None]:
    docs = sorted(
        [p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in SUPPORTED_DOC_EXTS],
        key=lambda p: p.name.lower(),
    )
    mandate = next((d for d in docs if any(k in d.name.lower() for k in MANDATE_KEYS)), None)
    bank = next((d for d in docs if any(k in d.name.lower() for k in BANK_KEYS)), None)
    # Fallback for mandate: prefer PDFs that do NOT match bank keywords
    if not mandate:
        non_bank = [d for d in docs if not any(k in d.name.lower() for k in BANK_KEYS)]
        mandate = next((d for d in non_bank if d.suffix.lower() == ".pdf"), None)
        # Last resort: any PDF that isn't already chosen as bank
        if not mandate:
            mandate = next((d for d in docs if d.suffix.lower() == ".pdf" and d != bank), None)
    if not bank:
        bank = next((d for d in docs if d != mandate), None)
    return mandate, bank


def brand_to_domain(brand: str) -> str:
    slug = brand.lower().replace(" ", "").strip()
    return f"www.{slug}.co.za"


# Map spreadsheet brand values (uppercase/variant) to HubSpot enum labels
BRAND_MAP: dict[str, str] = {
    "wimpy":           "Wimpy",
    "steers":          "Steers",
    "debonairs":       "Debonairs Pizza",
    "debonairs pizza": "Debonairs Pizza",
    "fishaways":       "Fishaways",
    "milky lane":      "Milky Lane",
    "mugg & bean":     "Mugg & Bean",
    "mugg and bean":   "Mugg & Bean",
    "netcafe":         "Netcafé",
    "netcafé":         "Netcafé",
    "paul":            "PAUL",
    "vovo telo":       "Vovo Telo",
    "fego caffe":      "Fego Caffé",
    "fego caffé":      "Fego Caffé",
    "turn 'n tender":  "Turn 'n Tender",
    "turn n tender":   "Turn 'n Tender",
    "mythos":          "Mythos",
    "salsa mexican grill": "Salsa Mexican Grill",
    "salsa":           "Salsa Mexican Grill",
    "lupa osteria":    "Lupa Osteria",
    "lupa":            "Lupa Osteria",
    "coffee couture":  "Coffee Couture",
    "famous brands multi": "Famous brands Multi",
}


def normalize_brand(brand: str) -> str:
    """Convert spreadsheet brand value to exact HubSpot enum label."""
    key = brand.strip().lower()
    return BRAND_MAP.get(key, brand.strip())  # fall back to original if unknown


def safe(val: object, fallback: str = "Undefined") -> str:
    s = str(val).strip() if val is not None else ""
    return s if s else fallback


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    if not TOKEN:
        print("ERROR: HUBSPOT_API_TOKEN is not set.")
        print("Set it first, for example:")
        print("  set HUBSPOT_API_TOKEN=pat-eu1-...")
        sys.exit(1)

    print("\n" + "=" * 60)
    print("  HubSpot Push - Test Run (first PASS/PASS row)")
    print("=" * 60)

    # ── Step 1: Load first PASS row ──────────────────────────────
    print("\n[1/9] Loading first PASS/PASS row from output workbook...")
    row = get_first_pass_row()

    subfolder      = safe(row.get("source subfolder"))
    company_name   = safe(row.get("company reg name"))
    reg_no         = safe(row.get("company registration number"))
    contact_email  = safe(row.get("primary franchisee email address"))
    first_name     = safe(row.get("Contact Name"))
    mobile         = safe(row.get("Contact Number") or row.get("Mobile Number"))
    fis_number     = safe(row.get("FIS Number"))
    restaurant_name = safe(row.get("Restaurant Name") or row.get("company reg name"))
    brand          = safe(row.get("Brand"))
    vat            = safe(row.get("VAT Number"))
    hs_brand       = normalize_brand(brand)

    print(f"       Subfolder:    {subfolder}")
    print(f"       Company:      {company_name}")
    print(f"       Registration: {reg_no}")
    print(f"       Contact:      {first_name} <{contact_email}>")
    print(f"       Mobile:       {mobile}")
    print(f"       Restaurant:   {restaurant_name}")
    print(f"       Brand:        {brand} -> {hs_brand}")
    print(f"       FIS:          {fis_number}")
    print(f"       VAT:          {vat}")

    # ── Step 2: Resolve files ────────────────────────────────────
    print(f"\n[2/9] Resolving document files from: {ROOT / subfolder}")
    folder_path = ROOT / subfolder
    if not folder_path.is_dir():
        print(f"       ERROR: Subfolder not found: {folder_path}")
        sys.exit(1)
    mandate_file, bank_file = classify_folder_docs(folder_path)
    print(f"       Mandate:    {mandate_file.name if mandate_file else 'NOT FOUND'}")
    print(f"       Bank proof: {bank_file.name if bank_file else 'NOT FOUND'}")

    # ── Step 3: Contact ──────────────────────────────────────────
    print(f"\n[3/9] Contact - searching by email: {contact_email}")
    contact_id = find_contact_by_email(contact_email)
    if contact_id:
        print(f"       Existing contact reused: id={contact_id}")
    else:
        print(f"       Not found - creating contact...")
        resp = api_post("/crm/v3/objects/contacts", {"properties": {
            "firstname":  first_name,
            "lastname":   "Undefined",
            "email":      contact_email,
            "mobilephone": mobile,
            "company_registration_no": reg_no,
        }})
        contact_id = resp.get("id")
        if contact_id:
            print(f"       Created contact: id={contact_id}")
        else:
            print("       ERROR: Contact creation failed. Aborting.")
            sys.exit(1)

    # ── Step 4: Company ──────────────────────────────────────────
    print(f"\n[4/9] Company - searching by registration: {reg_no}")
    company_id = find_company_by_registration(reg_no)
    if company_id:
        print(f"       Existing company reused: id={company_id}")
    else:
        print(f"       Not found - creating company...")
        resp = api_post("/crm/v3/objects/companies", {"properties": {
            "name":                     company_name,
            "domain":                   brand_to_domain(brand),
            "company_registration_no_": reg_no,
            "vat_number":               vat,
            "brand__famous_brands_":    hs_brand,
            "phone":                    mobile,
        }})
        company_id = resp.get("id")
        if company_id:
            print(f"       Created company: id={company_id}")
        else:
            print("       ERROR: Company creation failed. Aborting.")
            sys.exit(1)

    # ── Step 5: Associate contact -> company ─────────────────────
    print(f"\n[5/9] Associating contact {contact_id} -> company {company_id}")
    ok = associate_objects("contacts", contact_id, "companies", company_id, ASSOC_CONTACT_TO_COMPANY)
    print(f"       Result: {'OK' if ok else 'FAILED (non-fatal, continuing)'}")

    # ── Step 6: Upload mandate ───────────────────────────────────
    print(f"\n[6/9] Uploading mandate file...")
    mandate_url: str | None = None
    if mandate_file:
        mandate_url = upload_file(mandate_file)
        if not mandate_url:
            print("       File upload skipped (likely missing 'files' scope - add it in HubSpot private app settings)")
    else:
        print("       Skipped (no mandate file resolved)")

    # ── Step 7: Upload bank proof ────────────────────────────────
    print(f"\n[7/9] Uploading bank proof file...")
    bank_url: str | None = None
    if bank_file:
        bank_url = upload_file(bank_file)
        if not bank_url:
            print("       File upload skipped (likely missing 'files' scope - add it in HubSpot private app settings)")
    else:
        print("       Skipped (no bank proof file resolved)")

    # ── Step 8: Create deal ──────────────────────────────────────
    print(f"\n[8/9] Creating deal: '{restaurant_name}'")
    deal_props: dict[str, str] = {
        "dealname":               restaurant_name,
        "pipeline":               DEAL_PIPELINE,
        "dealstage":              DEAL_STAGE,
        "hubspot_owner_id":       DEAL_OWNER_ID,
        "fis_number":             fis_number,
        "fis__restaurant__number": fis_number,
        "nazmeys_store":          "true",
        "solution_required":      "eComm",
        "primary_acquiring_bank": "Nedbank",
        "request_type_new":       "New Store",
        "store_name":             restaurant_name,
        "brand__famous_brands_":  hs_brand,
        "vat_number":             vat,
    }
    if mandate_url:
        deal_props["debit_order_mandate"] = mandate_url
    if bank_url:
        deal_props["bank_letter"] = bank_url

    deal_resp = api_post("/crm/v3/objects/deals", {"properties": deal_props})
    deal_id = deal_resp.get("id")
    if deal_id:
        print(f"       Created deal: id={deal_id}")
    else:
        print("       ERROR: Deal creation failed. Aborting.")
        sys.exit(1)

    # ── Step 9: Associate deal -> company ────────────────────────
    print(f"\n[9/9] Associating deal {deal_id} -> company {company_id}")
    ok = associate_objects("deals", deal_id, "companies", company_id, ASSOC_DEAL_TO_COMPANY)
    print(f"       Result: {'OK' if ok else 'FAILED (non-fatal)'}")

    # ── Summary ──────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("  DONE")
    print(f"  Contact ID:  {contact_id}")
    print(f"  Company ID:  {company_id}")
    print(f"  Deal ID:     {deal_id}")
    print(f"  HubSpot URL: https://app.hubspot.com/contacts/{PORTAL_ID}/deal/{deal_id}")
    print("=" * 60)


if __name__ == "__main__":
    main()

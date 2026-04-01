import openpyxl
p = r'c:\Vibes\Hubpush\output v2 all fields.xlsx'
wb=openpyxl.load_workbook(p,data_only=True)
ws=wb.active
print('MAX_COL', ws.max_column)
print('HEADERS', [ws.cell(1,c).value for c in range(1, ws.max_column+1)])
for r in range(2,5):
    vals=[ws.cell(r,c).value for c in range(1, min(ws.max_column,14)+1)]
    print('ROW', r, vals)
wb.close()

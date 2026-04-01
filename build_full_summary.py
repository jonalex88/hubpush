from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

INPUT_PATH = Path(r"c:\Vibes\Hubpush\output v2 all fields FULL.resume.xlsx")
OUTPUT_PATH = Path(r"c:\Vibes\Hubpush\output v2 all fields FULL.summary.xlsx")

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")
PASS_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
PASS_FONT = Font(color="276221", bold=True)
FAIL_FONT = Font(color="9C0006", bold=True)


def style_header(ws, row=1):
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_pass_fail(cell):
    if cell.value == "PASS":
        cell.fill = PASS_FILL
        cell.font = PASS_FONT
    elif cell.value == "FAIL":
        cell.fill = FAIL_FILL
        cell.font = FAIL_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")


def main() -> None:
    wb = openpyxl.load_workbook(INPUT_PATH, data_only=True)
    ws = wb.active

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx = {str(h): i + 1 for i, h in enumerate(headers) if h is not None}

    source_subfolder_col = idx["source subfolder"]
    company_col = idx["company reg name"]
    mandate_col = idx["debit order mandate validation result"]
    bank_col = idx["bank account proof validation"]

    rows = []
    for r in range(2, ws.max_row + 1):
        company = ws.cell(r, company_col).value
        if company in (None, ""):
            continue
        rows.append(
            {
                "subfolder": str(ws.cell(r, source_subfolder_col).value or "").strip(),
                "company": str(company).strip(),
                "mandate": str(ws.cell(r, mandate_col).value or "").strip(),
                "bank": str(ws.cell(r, bank_col).value or "").strip(),
            }
        )

    total_rows = len(rows)
    mandate_pass = sum(1 for r in rows if r["mandate"] == "PASS")
    mandate_fail = sum(1 for r in rows if r["mandate"] == "FAIL")
    bank_pass = sum(1 for r in rows if r["bank"] == "PASS")
    bank_fail = sum(1 for r in rows if r["bank"] == "FAIL")
    both_pass = sum(1 for r in rows if r["mandate"] == "PASS" and r["bank"] == "PASS")
    any_fail = sum(1 for r in rows if r["mandate"] == "FAIL" or r["bank"] == "FAIL")
    unique_subfolders = len({r["subfolder"] for r in rows if r["subfolder"]})
    unique_companies = len({r["company"] for r in rows if r["company"]})

    summary_wb = openpyxl.Workbook()
    summary_ws = summary_wb.active
    summary_ws.title = "Summary"

    summary_ws.append(["Metric", "Value"])
    summary_rows = [
        ["Total output rows", total_rows],
        ["Unique subfolders represented", unique_subfolders],
        ["Unique companies represented", unique_companies],
        ["Mandate PASS", mandate_pass],
        ["Mandate FAIL", mandate_fail],
        ["Bank proof PASS", bank_pass],
        ["Bank proof FAIL", bank_fail],
        ["Both validations PASS", both_pass],
        ["At least one validation FAIL", any_fail],
    ]
    for row in summary_rows:
        summary_ws.append(row)

    style_header(summary_ws)
    summary_ws.column_dimensions["A"].width = 34
    summary_ws.column_dimensions["B"].width = 18
    summary_ws.freeze_panes = "A2"

    sub_ws = summary_wb.create_sheet("By Subfolder")
    sub_ws.append([
        "source subfolder",
        "row count",
        "mandate pass",
        "mandate fail",
        "bank pass",
        "bank fail",
        "both pass",
    ])

    by_subfolder = {}
    for row in rows:
        sub = row["subfolder"] or "(blank)"
        bucket = by_subfolder.setdefault(
            sub,
            {"count": 0, "mandate_pass": 0, "mandate_fail": 0, "bank_pass": 0, "bank_fail": 0, "both_pass": 0},
        )
        bucket["count"] += 1
        bucket["mandate_pass"] += 1 if row["mandate"] == "PASS" else 0
        bucket["mandate_fail"] += 1 if row["mandate"] == "FAIL" else 0
        bucket["bank_pass"] += 1 if row["bank"] == "PASS" else 0
        bucket["bank_fail"] += 1 if row["bank"] == "FAIL" else 0
        bucket["both_pass"] += 1 if row["mandate"] == "PASS" and row["bank"] == "PASS" else 0

    for subfolder in sorted(by_subfolder):
        b = by_subfolder[subfolder]
        sub_ws.append([
            subfolder,
            b["count"],
            b["mandate_pass"],
            b["mandate_fail"],
            b["bank_pass"],
            b["bank_fail"],
            b["both_pass"],
        ])

    style_header(sub_ws)
    sub_ws.freeze_panes = "A2"
    for col in ["A", "B", "C", "D", "E", "F", "G"]:
        sub_ws.column_dimensions[col].width = 20 if col == "A" else 14

    detail_ws = summary_wb.create_sheet("Failed Rows")
    detail_ws.append([
        "source subfolder",
        "company reg name",
        "mandate result",
        "bank result",
    ])
    for row in rows:
        if row["mandate"] == "FAIL" or row["bank"] == "FAIL":
            detail_ws.append([row["subfolder"], row["company"], row["mandate"], row["bank"]])

    style_header(detail_ws)
    detail_ws.freeze_panes = "A2"
    detail_ws.column_dimensions["A"].width = 22
    detail_ws.column_dimensions["B"].width = 40
    detail_ws.column_dimensions["C"].width = 16
    detail_ws.column_dimensions["D"].width = 16
    for r in range(2, detail_ws.max_row + 1):
        style_pass_fail(detail_ws.cell(r, 3))
        style_pass_fail(detail_ws.cell(r, 4))

    summary_wb.save(OUTPUT_PATH)
    wb.close()
    summary_wb.close()
    print(f"Created summary workbook: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()

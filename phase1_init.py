from __future__ import annotations

import argparse
import shutil
from pathlib import Path

import openpyxl

from hubpush_core.data_model import (
    SYNC_COLUMNS,
    build_cloud_row_id,
    compute_row_checksum,
    default_hubspot_status,
    normalize_text,
    utc_now_iso,
)
from hubpush_core.sync_store import append_jsonl, read_json, write_json


def ensure_columns(ws) -> dict[str, int]:
    headers = [normalize_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    idx = {h: i + 1 for i, h in enumerate(headers) if h}

    for col_name in SYNC_COLUMNS:
        if col_name not in idx:
            ws.cell(1, ws.max_column + 1, col_name)
            idx[col_name] = ws.max_column

    # Rebuild index after possible inserts.
    headers = [normalize_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    return {h: i + 1 for i, h in enumerate(headers) if h}


def row_to_dict(ws, row_num: int, header_idx: dict[str, int]) -> dict[str, str]:
    out: dict[str, str] = {}
    for h, c in header_idx.items():
        out[h] = normalize_text(ws.cell(row_num, c).value)
    return out


def write_row(ws, row_num: int, header_idx: dict[str, int], key: str, value: str) -> None:
    col = header_idx.get(key)
    if col:
        ws.cell(row_num, col, value)


def init_master_spreadsheet(master_path: Path, snapshot_path: Path, commits_path: Path) -> dict[str, int]:
    backup = master_path.with_suffix(master_path.suffix + ".phase1.bak")
    shutil.copy2(master_path, backup)

    wb = openpyxl.load_workbook(master_path)
    ws = wb.active

    header_idx = ensure_columns(ws)

    changed_rows = 0
    id_counts: dict[str, int] = {}
    rows_out = []
    now_iso = utc_now_iso()

    for r in range(2, ws.max_row + 1):
        row = row_to_dict(ws, r, header_idx)
        company = row.get("company reg name", "")
        if not company:
            continue

        base_id = build_cloud_row_id(row)
        id_counts[base_id] = id_counts.get(base_id, 0) + 1
        row_id = base_id if id_counts[base_id] == 1 else f"{base_id}-{id_counts[base_id]}"
        if row.get("Cloud Row ID") != row_id:
            write_row(ws, r, header_idx, "Cloud Row ID", row_id)
            row["Cloud Row ID"] = row_id
            changed_rows += 1

        if not row.get("HubSpot Status"):
            status = default_hubspot_status(row)
            write_row(ws, r, header_idx, "HubSpot Status", status)
            row["HubSpot Status"] = status
            changed_rows += 1

        row["Cloud Updated At"] = now_iso
        write_row(ws, r, header_idx, "Cloud Updated At", now_iso)

        checksum = compute_row_checksum(row)
        write_row(ws, r, header_idx, "Row Checksum", checksum)
        row["Row Checksum"] = checksum

        rows_out.append(row)

    wb.save(master_path)
    wb.close()

    snapshot = {
        "schema_version": 1,
        "master_file": str(master_path),
        "updated_at": now_iso,
        "row_count": len(rows_out),
        "rows": rows_out,
    }
    write_json(snapshot_path, snapshot)

    commits_meta = read_json(commits_path, {"schema_version": 1, "commits": []})
    write_json(commits_path, commits_meta)

    append_jsonl(
        commits_path.with_suffix(".events.jsonl"),
        {
            "event": "phase1_init",
            "updated_at": now_iso,
            "rows_indexed": len(rows_out),
            "rows_changed": changed_rows,
            "backup_file": str(backup),
        },
    )

    return {
        "rows_indexed": len(rows_out),
        "rows_changed": changed_rows,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Phase 1 initializer for HubPush data model")
    parser.add_argument(
        "--master",
        default=r"c:\Vibes\Hubpush\output v2 all fields FULL.resume.xlsx",
        help="Path to the master spreadsheet",
    )
    parser.add_argument(
        "--snapshot",
        default=r"c:\Vibes\Hubpush\data\cloud_master_snapshot.json",
        help="Path to write the cloud snapshot JSON",
    )
    parser.add_argument(
        "--commits",
        default=r"c:\Vibes\Hubpush\data\commit_history.json",
        help="Path to write commit history JSON",
    )
    args = parser.parse_args()

    master_path = Path(args.master)
    snapshot_path = Path(args.snapshot)
    commits_path = Path(args.commits)

    if not master_path.exists():
        raise FileNotFoundError(f"Master spreadsheet not found: {master_path}")

    result = init_master_spreadsheet(master_path, snapshot_path, commits_path)
    print("Phase 1 initialized")
    print(f"Rows indexed: {result['rows_indexed']}")
    print(f"Rows changed: {result['rows_changed']}")
    print(f"Master: {master_path}")
    print(f"Snapshot: {snapshot_path}")
    print(f"Commit store: {commits_path}")


if __name__ == "__main__":
    main()

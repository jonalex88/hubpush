"""
Phase 3: HubSpot Service Module - Usage Examples

This script demonstrates how to use hubpush_core.hubspot_service for pushing
restaurant rows to HubSpot with full contact/company/deal orchestration.
"""

import logging
from pathlib import Path
from hubpush_core.hubspot_service import HubSpotConfig, HubSpotService
from hubpush_core.data_model import build_cloud_row_id, compute_row_checksum

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

# ── EXAMPLE 1: Single Row Push (Dry Run) ──────────────────────────────────

def example_dry_run():
    """
    Test push without actually calling HubSpot API.
    Useful for validating row data and document classification.
    """
    logger.info("\n" + "=" * 60)
    logger.info("EXAMPLE 1: DRY-RUN PUSH")
    logger.info("=" * 60)
    
    # Initialize service in dry-run mode
    config = HubSpotConfig(
        token="pat-eu1-YOUR-TOKEN-HERE",
        dry_run=True,  # No actual API calls
    )
    hs = HubSpotService(config)
    
    # Example row from spreadsheet
    row = {
        "source subfolder": "Wimpy/0001",
        "company reg name": "Wimpy South Africa (Pty) Ltd",
        "company registration number": "2001234567",
        "primary franchisee email address": "franchisee@example.com",
        "Contact Name": "John Doe",
        "Contact Number": "+27111234567",
        "FIS Number": "12345",
        "Restaurant Name": "Wimpy - Sandton",
        "Brand": "Wimpy",
        "VAT Number": "4123456789",
    }
    
    file_root = Path(r"c:\Vibes\Hubpush\1.FBEO\1.FBEO")
    
    result = hs.push_row(row, file_root, dry_run=True)
    
    logger.info(f"\nResult:")
    logger.info(f"  OK: {result.ok}")
    logger.info(f"  Contact ID: {result.contact_id}")
    logger.info(f"  Company ID: {result.company_id}")
    logger.info(f"  Deal ID: {result.deal_id}")
    logger.info(f"  Error: {result.error}")


# ── EXAMPLE 2: Batch Push Multiple Rows ──────────────────────────────────

def example_batch_push(rows: list[dict], dry_run: bool = True):
    """
    Push multiple rows to HubSpot, collecting results.
    """
    logger.info("\n" + "=" * 60)
    logger.info(f"BATCH PUSH: {len(rows)} rows")
    logger.info("=" * 60)
    
    config = HubSpotConfig(
        token="pat-eu1-YOUR-TOKEN-HERE",
        dry_run=dry_run,
    )
    hs = HubSpotService(config)
    file_root = Path(r"c:\Vibes\Hubpush\1.FBEO\1.FBEO")
    
    results = []
    success_count = 0
    failed_count = 0
    
    for i, row in enumerate(rows, 1):
        logger.info(f"\n[{i}/{len(rows)}] Pushing {row.get('Restaurant Name', 'Unknown')}")
        result = hs.push_row(row, file_root)
        results.append(result)
        
        if result.ok:
            success_count += 1
            logger.info(f"  ✓ Success: contact={result.contact_id}, company={result.company_id}, deal={result.deal_id}")
        else:
            failed_count += 1
            logger.error(f"  ✗ Failed: {result.error}")
    
    logger.info(f"\n{'=' * 60}")
    logger.info(f"Summary: {success_count} succeeded, {failed_count} failed")
    logger.info(f"{'=' * 60}")
    
    return results


# ── EXAMPLE 3: Load Rows from Excel ──────────────────────────────────────

def example_from_excel():
    """
    Load rows from master spreadsheet and push "Ready To Push" rows.
    """
    import openpyxl
    from hubpush_core.data_model import default_hubspot_status, STATUS_READY_TO_PUSH
    
    logger.info("\n" + "=" * 60)
    logger.info("EXAMPLE 3: LOAD FROM EXCEL AND PUSH")
    logger.info("=" * 60)
    
    excel_file = Path(r"c:\Vibes\Hubpush\output v2 all fields FULL.resume.xlsx")
    file_root = Path(r"c:\Vibes\Hubpush\1.FBEO\1.FBEO")
    
    # Load rows
    wb = openpyxl.load_workbook(str(excel_file), data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    
    ready_rows = []
    for r in range(2, min(10, ws.max_row + 1)):  # First 9 rows for demo
        row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        status = default_hubspot_status(row)
        
        if status == STATUS_READY_TO_PUSH:
            ready_rows.append(row)
            logger.info(f"  Found ready: {row.get('Restaurant Name', 'Unknown')}")
    
    wb.close()
    
    logger.info(f"\nFound {len(ready_rows)} ready rows to push.")
    
    # Dry-run push
    results = example_batch_push(ready_rows, dry_run=True)
    
    return results


# ── EXAMPLE 4: Brand Normalization ──────────────────────────────────────

def example_brand_normalization():
    """
    Demonstrate brand name mapping to HubSpot enum values.
    """
    logger.info("\n" + "=" * 60)
    logger.info("EXAMPLE 4: BRAND NORMALIZATION")
    logger.info("=" * 60)
    
    config = HubSpotConfig(token="dummy")
    hs = HubSpotService(config)
    
    test_brands = [
        "wimpy", "WIMPY", "Wimpy",
        "debonairs", "Debonairs Pizza",
        "mugg & bean",
        "unknown brand",
    ]
    
    for brand in test_brands:
        normalized = hs.normalize_brand(brand)
        domain = hs.brand_to_domain(brand)
        logger.info(f"  {brand:20} -> {normalized:25} ({domain})")


# ── EXAMPLE 5: File Classification ──────────────────────────────────────

def example_file_classification():
    """
    Demonstrate automatic document classification (mandate vs bank proof).
    """
    logger.info("\n" + "=" * 60)
    logger.info("EXAMPLE 5: FILE CLASSIFICATION")
    logger.info("=" * 60)
    
    # Create test folder structure
    test_folder = Path(r"c:\Vibes\Hubpush\1.FBEO\1.FBEO\Wimpy\0001")
    
    if test_folder.exists():
        mandate, bank = HubSpotService.classify_folder_docs(test_folder)
        logger.info(f"  Folder: {test_folder}")
        logger.info(f"  Mandate: {mandate.name if mandate else 'NOT FOUND'}")
        logger.info(f"  Bank proof: {bank.name if bank else 'NOT FOUND'}")
    else:
        logger.warning(f"  Test folder not found: {test_folder}")


# ── EXAMPLE 6: Data Model Integration ──────────────────────────────────────

def example_data_model_integration():
    """
    Show how HubSpot service integrates with data model (Cloud Row ID, checksums).
    """
    logger.info("\n" + "=" * 60)
    logger.info("EXAMPLE 6: DATA MODEL INTEGRATION")
    logger.info("=" * 60)
    
    row = {
        "source subfolder": "Wimpy/0001",
        "company reg name": "Wimpy South Africa (Pty) Ltd",
        "company registration number": "2001234567",
        "primary franchisee email address": "franchisee@example.com",
        "FIS Number": "12345",
        "Restaurant Name": "Wimpy - Sandton",
        "Email Address": "info@wimpy.com",
        "Contact Email": "",
    }
    
    # Generate deterministic Cloud Row ID
    cloud_row_id = build_cloud_row_id(row)
    logger.info(f"  Cloud Row ID: {cloud_row_id}")
    
    # Compute row checksum (for drift detection)
    checksum = compute_row_checksum(row)
    logger.info(f"  Row Checksum: {checksum}")
    
    # These IDs ensure that the same row pushed twice won't create duplicates
    # (HubSpot service uses email + registration lookup to reuse existing contacts/companies)


# ── MAIN ──────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    logger.info("HubSpot Service Module - Usage Examples\n")
    
    # Run examples
    example_dry_run()
    example_brand_normalization()
    example_file_classification()
    example_data_model_integration()
    # example_from_excel()  # Uncomment to test with real Excel data
    
    logger.info("\n" + "=" * 60)
    logger.info("Examples complete!")
    logger.info("=" * 60)

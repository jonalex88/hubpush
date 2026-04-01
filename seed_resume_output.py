import openpyxl
from pathlib import Path

full = Path(r'c:\Vibes\Hubpush\output v2 all fields FULL.xlsx')
out = Path(r'c:\Vibes\Hubpush\output v2 all fields FULL.resume.xlsx')
if not full.exists():
    print('NO_SOURCE_FILE')
else:
    wb = openpyxl.load_workbook(full, data_only=True)
    ws = wb.active
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = ws.title
    for c in range(1, ws.max_column + 1):
        ws2.cell(1, c, ws.cell(1, c).value)
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws2.cell(r, c, ws.cell(r, c).value)
    wb2.save(out)
    wb.close()
    wb2.close()
    print('SEEDED', out)

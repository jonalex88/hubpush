import argparse
import re
import sys
import time
from datetime import datetime, date
from pathlib import Path
from typing import Iterable

import fitz
import openpyxl
import pytesseract
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image
from pypdf import PdfReader

SUPPORTED_DOC_EXTS = {".pdf", ".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp", ".webp"}
MANDATE_KEYS = ("mandate", "debit", "order")
BANK_KEYS = ("bank", "account", "confirmation", "proof", "statement")

DATE_PATTERNS = [
    "%Y-%m-%d",
    "%d-%m-%Y",
    "%d/%m/%Y",
    "%d.%m.%Y",
    "%d %b %Y",
    "%d %B %Y",
    "%b %d %Y",
    "%B %d %Y",
]

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")
PASS_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
PASS_FONT = Font(color="276221", bold=True)
FAIL_FONT = Font(color="9C0006", bold=True)
HEARTBEAT_INTERVAL_SECONDS = 20


def heartbeat(message: str, force: bool = False) -> None:
    now = time.time()
    last = getattr(heartbeat, "_last", 0.0)
    if force or (now - last) >= HEARTBEAT_INTERVAL_SECONDS:
        print(f"HEARTBEAT {datetime.now().strftime('%H:%M:%S')} | {message}")
        sys.stdout.flush()
        heartbeat._last = now


def normalize(s: str) -> str:
    return re.sub(r"\s+", " ", s.strip().lower())


def company_tokens(company_name: str) -> set[str]:
    words = re.findall(r"[a-z0-9]+", company_name.lower())
    stop = {
        "pty",
        "ltd",
        "limited",
        "cc",
        "the",
        "and",
        "of",
        "sa",
        "south",
        "africa",
    }
    return {w for w in words if len(w) >= 3 and w not in stop}


def configure_tesseract() -> None:
    candidates = [
        Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe"),
        Path(r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe"),
    ]
    for c in candidates:
        if c.exists():
            pytesseract.pytesseract.tesseract_cmd = str(c)
            return


def clean_registration_number(value: object) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    return re.sub(r"\s+", "", s)


def stringify_cell_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def parse_excel_rows(xlsx_path: Path) -> tuple[list[dict], list[str]]:
    rows: list[dict] = []
    discovered_headers: list[str] = []
    seen_headers: set[str] = set()

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as exc:
        print(f"WARNING: Could not read spreadsheet {xlsx_path.name}: {exc}")
        return rows, discovered_headers

    try:
        for ws in wb.worksheets:
            # Skip sample/template sheets when present.
            if "example" in ws.title.lower():
                continue

            max_row = min(ws.max_row, 2000)
            max_col = min(ws.max_column, 120)
            header_row = None
            company_col = None
            header_map: dict[int, str] = {}

            for r in range(1, min(max_row, 200) + 1):
                for c in range(1, max_col + 1):
                    val = ws.cell(r, c).value
                    if val is None:
                        continue
                    txt = normalize(str(val))
                    if "registered company name" in txt:
                        header_row = r
                        company_col = c
                if header_row is not None:
                    break

            if header_row is None or company_col is None:
                continue

            for c in range(1, max_col + 1):
                header_val = ws.cell(header_row, c).value
                if header_val is None:
                    continue
                header_name = str(header_val).strip()
                if not header_name:
                    continue
                header_map[c] = header_name
                if header_name not in seen_headers:
                    seen_headers.add(header_name)
                    discovered_headers.append(header_name)

            for r in range(header_row + 1, max_row + 1):
                cell = ws.cell(r, company_col).value
                if cell is None:
                    continue
                name = str(cell).strip()
                if not name:
                    continue
                if normalize(name) in {"none", "n/a", "na"}:
                    continue

                source_fields: dict[str, str] = {}
                for c, h in header_map.items():
                    source_fields[h] = stringify_cell_value(ws.cell(r, c).value)

                rows.append(
                    {
                        "company": name,
                        "source_fields": source_fields,
                    }
                )
    finally:
        wb.close()

    return rows, discovered_headers


def build_customer_failure_paragraph(mandate_result: str, bank_result: str, failures: list[str]) -> str:
    if mandate_result == "PASS" and bank_result == "PASS":
        return "Documentation validation is complete and all submitted records meet onboarding requirements. No further action is required at this stage."

    issues = []
    actions = []

    text = " ".join(failures).lower()

    if "missing debit order mandate file" in text:
        issues.append("we did not receive a debit order mandate")
        actions.append("submit a signed debit order mandate for this registered company")
    elif "not confidently identified as signed" in text:
        issues.append("the submitted debit order mandate could not be verified as signed")
        actions.append("resubmit a clearly signed debit order mandate")

    if "no bank account number found in mandate" in text:
        issues.append("the bank account number could not be read from the debit order mandate")
        actions.append("provide a readable mandate showing the bank account number")

    if "missing proof of bank account file" in text:
        issues.append("we did not receive proof of bank account")
        actions.append("submit an official proof of bank account document from your bank")

    if "older than january 2025" in text:
        issues.append("the proof of bank account is dated before January 2025")
        actions.append("provide a proof of bank account dated January 2025 or later")

    if "account mismatch" in text:
        issues.append("the bank account number on the proof of bank account does not match the mandate")
        actions.append("submit matching documents that reflect the same bank account number")
    elif "could not confirm account number match" in text:
        issues.append("the account number match could not be confirmed")
        actions.append("provide clearer documents where the account number is clearly visible on both documents")

    if not issues:
        issues.append("some required validation checks could not be completed")
    if not actions:
        actions.append("resubmit clear and complete onboarding documents for this company")

    issues_sentence = "; ".join(dict.fromkeys(issues))
    actions_sentence = "; ".join(dict.fromkeys(actions))

    return (
        "During documentation validation, we identified the following issue(s): "
        f"{issues_sentence}. "
        "To finalise onboarding, please "
        f"{actions_sentence}."
    )


def list_documents(folder: Path) -> list[Path]:
    docs = [p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in SUPPORTED_DOC_EXTS]
    return sorted(docs, key=lambda p: p.name.lower())


def score_filename_match(file_path: Path, company_name: str) -> int:
    fname = file_path.stem.lower()
    score = 0
    for tok in company_tokens(company_name):
        if tok in fname:
            score += 2
    return score


def pick_document_for_company(company_name: str, candidates: list[Path]) -> Path | None:
    if not candidates:
        return None
    scored = sorted(candidates, key=lambda p: score_filename_match(p, company_name), reverse=True)
    if score_filename_match(scored[0], company_name) > 0:
        return scored[0]
    return candidates[0]


def extract_text_pdf_native(pdf_path: Path) -> str:
    out = []
    try:
        reader = PdfReader(str(pdf_path))
        for page in reader.pages[:5]:
            out.append(page.extract_text() or "")
    except Exception:
        return ""
    return "\n".join(out)


def extract_text_ocr_image(image: Image.Image) -> str:
    gray = image.convert("L")
    return pytesseract.image_to_string(gray)


def extract_text_from_file(path: Path) -> str:
    if path.suffix.lower() == ".pdf":
        try:
            heartbeat(f"Reading PDF: {path.name}")
            native = extract_text_pdf_native(path)
            if len(native.strip()) >= 80:
                return native

            ocr_text = []
            doc = fitz.open(path)
            try:
                pages = min(doc.page_count, 4)
                for i in range(pages):
                    heartbeat(f"OCR page {i + 1}/{pages}: {path.name}")
                    try:
                        pix = doc.load_page(i).get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
                        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                        ocr_text.append(extract_text_ocr_image(img))
                    except Exception:
                        continue
            finally:
                doc.close()
            return "\n".join(ocr_text)
        except Exception:
            return ""

    try:
        heartbeat(f"OCR image: {path.name}")
        with Image.open(path) as im:
            return extract_text_ocr_image(im)
    except Exception:
        return ""


def document_read_error(path: Path) -> str | None:
    try:
        if path.suffix.lower() == ".pdf":
            try:
                _ = PdfReader(str(path))
                return None
            except Exception:
                try:
                    doc = fitz.open(path)
                    doc.close()
                    return None
                except Exception as exc:
                    return f"{exc.__class__.__name__}: {exc}"

        with Image.open(path) as im:
            im.verify()
        return None
    except Exception as exc:
        return f"{exc.__class__.__name__}: {exc}"


def extract_pdf_has_digital_signature(pdf_path: Path) -> bool:
    try:
        reader = PdfReader(str(pdf_path))
        fields = reader.get_fields() or {}
        for _, field in fields.items():
            field_type = field.get("/FT")
            value = field.get("/V")
            if str(field_type) == "/Sig" and value is not None:
                return True
    except Exception:
        return False
    return False


def looks_signed_by_text(text: str) -> bool:
    low = text.lower()
    if "signature" not in low and "signed" not in low:
        return False

    # Capture patterns like: Signature: John Doe
    m = re.search(r"signature\s*[:\-]\s*([a-z0-9][a-z0-9 .,&'/-]{2,})", low)
    if m:
        value = m.group(1).strip()
        if value not in {"", "n/a", "none", "na", "nil"}:
            return True

    # If OCR catches explicit statement.
    if re.search(r"(signed by|electronically signed|authorised signature)", low):
        return True

    return False


def looks_signed_by_image(path: Path) -> bool:
    try:
        if path.suffix.lower() == ".pdf":
            doc = fitz.open(path)
            try:
                if doc.page_count == 0:
                    return False
                pix = doc.load_page(0).get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            finally:
                doc.close()
        else:
            img = Image.open(path)

        gray = img.convert("L")
        w, h = gray.size
        # Signature blocks are usually in lower half.
        crop = gray.crop((0, int(h * 0.55), w, h))
        pixels = crop.getdata()
        dark = sum(1 for p in pixels if p < 90)
        ratio = dark / max(len(pixels), 1)
        return ratio > 0.01
    except Exception:
        return False


def extract_account_candidates(text: str) -> list[str]:
    nums = re.findall(r"(?<!\d)(\d{8,16})(?!\d)", text)
    # Remove obvious date-like numbers and duplicates, preserve order.
    seen = set()
    result = []
    for n in nums:
        if n.startswith("20") and len(n) == 8:
            continue
        if n not in seen:
            seen.add(n)
            result.append(n)
    return result


def extract_mandate_account(text: str) -> str | None:
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    for ln in lines:
        low = ln.lower()
        if "account" in low and "number" in low:
            nums = extract_account_candidates(ln)
            if nums:
                return nums[0]

    nums = extract_account_candidates(text)
    return nums[0] if nums else None


def parse_possible_dates(text: str) -> list[date]:
    text = text.replace(",", " ")
    candidates = set()

    for m in re.findall(r"\b\d{4}[-/]\d{1,2}[-/]\d{1,2}\b", text):
        candidates.add(m.replace("/", "-"))

    for m in re.findall(r"\b\d{1,2}[./-]\d{1,2}[./-]\d{4}\b", text):
        candidates.add(m)

    for m in re.findall(r"\b\d{1,2}\s+[A-Za-z]{3,9}\s+\d{4}\b", text):
        candidates.add(m)

    for m in re.findall(r"\b[A-Za-z]{3,9}\s+\d{1,2}\s+\d{4}\b", text):
        candidates.add(m)

    dates = []
    for c in candidates:
        for fmt in DATE_PATTERNS:
            try:
                d = datetime.strptime(c, fmt).date()
                if 2000 <= d.year <= 2100:
                    dates.append(d)
                    break
            except ValueError:
                pass

    return sorted(set(dates))


def validate_mandate(mandate_file: Path | None) -> tuple[str, str | None, list[str]]:
    failures = []
    if mandate_file is None:
        return "FAIL", None, ["Missing debit order mandate file"]

    read_error = document_read_error(mandate_file)
    if read_error:
        return (
            "FAIL",
            None,
            [f"Mandate file could not be read or may be corrupt ({mandate_file.name}; {read_error})"],
        )

    text = extract_text_from_file(mandate_file)
    signed = False

    if mandate_file.suffix.lower() == ".pdf" and extract_pdf_has_digital_signature(mandate_file):
        signed = True
    if not signed and looks_signed_by_text(text):
        signed = True
    if not signed and looks_signed_by_image(mandate_file):
        # Fallback heuristic for scanned docs.
        signed = True

    if not signed:
        failures.append(f"Mandate not confidently identified as signed ({mandate_file.name})")

    account = extract_mandate_account(text)
    if not account:
        failures.append(f"No bank account number found in mandate ({mandate_file.name})")

    return ("PASS" if signed else "FAIL"), account, failures


def validate_bank_proof(bank_file: Path | None, mandate_account: str | None) -> tuple[str, list[str]]:
    failures = []
    if bank_file is None:
        return "FAIL", ["Missing proof of bank account file"]

    read_error = document_read_error(bank_file)
    if read_error:
        return (
            "FAIL",
            [f"Proof of bank account file could not be read or may be corrupt ({bank_file.name}; {read_error})"],
        )

    text = extract_text_from_file(bank_file)
    dates = parse_possible_dates(text)

    if not dates:
        failures.append(f"No clear document date found on bank proof ({bank_file.name})")
    else:
        picked = max(dates)
        if picked < date(2025, 1, 1):
            failures.append(
                f"Bank proof date is older than January 2025 ({picked.isoformat()} in {bank_file.name})"
            )

    bank_accounts = extract_account_candidates(text)
    bank_account = bank_accounts[0] if bank_accounts else None

    if mandate_account and bank_account:
        if mandate_account != bank_account:
            failures.append(
                f"Account mismatch: mandate {mandate_account} vs bank proof {bank_account}"
            )
    else:
        failures.append(
            f"Could not confirm account number match (mandate={mandate_account or 'not found'}, bank proof={bank_account or 'not found'})"
        )

    return ("PASS" if not failures else "FAIL"), failures


def classify_docs_for_folder(docs: Iterable[Path]) -> tuple[list[Path], list[Path]]:
    mandate = []
    bank = []
    for d in docs:
        low = d.name.lower()
        if any(k in low for k in MANDATE_KEYS):
            mandate.append(d)
        elif any(k in low for k in BANK_KEYS):
            bank.append(d)

    docs = list(docs)
    if not mandate:
        mandate = [d for d in docs if d.suffix.lower() == ".pdf"][:1] or docs[:1]
    if not bank:
        leftovers = [d for d in docs if d not in mandate]
        bank = leftovers[:1]

    return mandate, bank


def write_output(rows: list[dict], source_headers: list[str], output_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "outputv1"

    core_headers = [
        "source subfolder",
        "company reg name",
        "company registration number",
        "primary franchisee email address",
        "debit order mandate validation result",
        "bank account proof validation",
        "description of any validation failures",
    ]
    core_normalized = {normalize(h) for h in core_headers}
    appended_headers = [h for h in source_headers if normalize(h) not in core_normalized]
    headers = core_headers + appended_headers

    for idx, h in enumerate(headers, 1):
        c = ws.cell(1, idx, h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 38
    ws.column_dimensions["E"].width = 36
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 110
    for col_idx in range(8, len(headers) + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 24

    for r_idx, row in enumerate(rows, start=2):
        ws.cell(r_idx, 1, row["subfolder"]).alignment = Alignment(vertical="top")
        ws.cell(r_idx, 2, row["company"]).alignment = Alignment(vertical="top")
        ws.cell(r_idx, 3, row["registration_number"]).alignment = Alignment(vertical="top")
        ws.cell(r_idx, 4, row["primary_email"]).alignment = Alignment(vertical="top")

        mcell = ws.cell(r_idx, 5, row["mandate_result"])
        bcell = ws.cell(r_idx, 6, row["bank_result"])
        ws.cell(r_idx, 7, row["failures"]).alignment = Alignment(wrap_text=True, vertical="top")

        for cell in (mcell, bcell):
            if cell.value == "PASS":
                cell.fill = PASS_FILL
                cell.font = PASS_FONT
            else:
                cell.fill = FAIL_FILL
                cell.font = FAIL_FONT
            cell.alignment = Alignment(horizontal="center", vertical="top")

        source_fields = row.get("source_fields", {})
        for offset, header_name in enumerate(appended_headers, start=8):
            ws.cell(r_idx, offset, source_fields.get(header_name, "")).alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)


def load_existing_row_keys(output_path: Path) -> set[tuple[str, str, str, str]]:
    if not output_path.exists():
        return set()

    keys: set[tuple[str, str, str, str]] = set()
    try:
        wb = openpyxl.load_workbook(output_path, data_only=True)
    except Exception:
        return keys

    try:
        ws = wb.active
        headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
        idx = {h: i + 1 for i, h in enumerate(headers)}

        comp_col = idx.get("company reg name")
        reg_col = idx.get("company registration number")
        email_col = idx.get("primary franchisee email address")
        fis_col = idx.get("FIS Number")

        if not comp_col:
            return keys

        for r in range(2, ws.max_row + 1):
            comp = str(ws.cell(r, comp_col).value or "").strip()
            if not comp:
                continue
            reg = str(ws.cell(r, reg_col).value or "").strip() if reg_col else ""
            email = str(ws.cell(r, email_col).value or "").strip() if email_col else ""
            fis = str(ws.cell(r, fis_col).value or "").strip() if fis_col else ""
            keys.add((normalize(comp), normalize(reg), normalize(email), normalize(fis)))
    finally:
        wb.close()

    return keys


def run(root: Path, output: Path) -> None:
    configure_tesseract()

    all_rows = []
    all_source_headers: list[str] = []
    seen_source_headers: set[str] = set()

    existing_keys = load_existing_row_keys(output)
    if existing_keys:
        print(f"Resume mode: found {len(existing_keys)} existing output rows to avoid reprocessing.")

    subfolders = sorted([p for p in root.iterdir() if p.is_dir()], key=lambda p: p.name.lower())
    total_subfolders = len(subfolders)
    processed_subfolders = 0
    skipped_subfolders = 0

    for idx_sub, sub in enumerate(subfolders, start=1):
        heartbeat(f"Folder {idx_sub}/{total_subfolders}: {sub.name}", force=True)
        xlsx_files = sorted(sub.glob("*.xlsx"))
        if not xlsx_files:
            print(f"[{idx_sub}/{total_subfolders}] Skip {sub.name} (no spreadsheet).")
            continue

        folder_records: list[dict] = []
        for xf in xlsx_files:
            parsed_rows, parsed_headers = parse_excel_rows(xf)
            folder_records.extend(parsed_rows)
            for h in parsed_headers:
                if h not in seen_source_headers:
                    seen_source_headers.add(h)
                    all_source_headers.append(h)

        if not folder_records:
            print(f"[{idx_sub}/{total_subfolders}] Skip {sub.name} (no company records found).")
            continue

        folder_keys = set()
        for rec in folder_records:
            fields = rec.get("source_fields", {})
            folder_keys.add(
                (
                    normalize(rec.get("company", "")),
                    normalize(clean_registration_number(fields.get("Company Registration Number", ""))),
                    normalize(str(fields.get("Primary User Email Address", "")).strip()),
                    normalize(str(fields.get("FIS Number", "")).strip()),
                )
            )

        if existing_keys and folder_keys and folder_keys.issubset(existing_keys):
            skipped_subfolders += 1
            print(f"[{idx_sub}/{total_subfolders}] Skip {sub.name} (already in existing output).")
            continue

        companies: dict[str, dict[str, str]] = {}
        for rec in folder_records:
            cname = rec["company"]
            fields = rec.get("source_fields", {})
            reg_no = clean_registration_number(fields.get("Company Registration Number", ""))
            primary_email = fields.get("Primary User Email Address", "").strip()

            existing = companies.get(cname)
            if existing:
                if not existing.get("registration_number") and reg_no:
                    existing["registration_number"] = reg_no
                if not existing.get("primary_email") and primary_email:
                    existing["primary_email"] = primary_email
            else:
                companies[cname] = {
                    "registration_number": reg_no,
                    "primary_email": primary_email,
                }

        docs = list_documents(sub)
        mandate_candidates, bank_candidates = classify_docs_for_folder(docs)

        company_validation: dict[str, dict[str, str]] = {}

        for company in sorted(companies):
            mandate_file = pick_document_for_company(company, mandate_candidates)
            bank_file = pick_document_for_company(company, bank_candidates)

            mandate_result, mandate_account, mandate_fails = validate_mandate(mandate_file)
            bank_result, bank_fails = validate_bank_proof(bank_file, mandate_account)

            all_failures = mandate_fails + bank_fails
            paragraph = build_customer_failure_paragraph(mandate_result, bank_result, all_failures)

            company_validation[company] = {
                "mandate_result": mandate_result,
                "bank_result": bank_result,
                "failures": paragraph,
            }

        for rec in folder_records:
            company = rec["company"]
            validation = company_validation.get(
                company,
                {
                    "mandate_result": "FAIL",
                    "bank_result": "FAIL",
                    "failures": "During documentation validation, we identified the following issue(s): some required validation checks could not be completed. To finalise onboarding, please resubmit clear and complete onboarding documents for this company.",
                },
            )
            all_rows.append(
                {
                    "subfolder": sub.name,
                    "company": company,
                    "registration_number": companies.get(company, {}).get("registration_number", ""),
                    "primary_email": companies.get(company, {}).get("primary_email", ""),
                    "mandate_result": validation["mandate_result"],
                    "bank_result": validation["bank_result"],
                    "failures": validation["failures"],
                    "source_fields": rec.get("source_fields", {}),
                }
            )

        processed_subfolders += 1
        print(
            f"[{idx_sub}/{total_subfolders}] Processed {sub.name} | "
            f"new rows: {len(folder_records)} | processed: {processed_subfolders} | skipped: {skipped_subfolders}"
        )
        sys.stdout.flush()

    write_output(all_rows, all_source_headers, output)
    print(f"Wrote {len(all_rows)} rows to {output}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", required=True, help="Root folder containing subfolders")
    parser.add_argument("--output", required=True, help="Output xlsx path")
    args = parser.parse_args()

    run(Path(args.root), Path(args.output))


if __name__ == "__main__":
    main()

"""
Document Validation Tool — Company-Aware Version
Reads each subfolder's signup Excel, extracts registered company names,
then validates the debit order mandate and proof of bank account for each company.

Usage:
    python validate_documents.py --folder "C:/path/to/folders" --output outputv1.xlsx
    python validate_documents.py --folder "C:/Vibes/Hubpush/1.FBEO/Sample" --output "C:/Vibes/Hubpush/outputv1.xlsx"

Requirements:
    pip install anthropic openpyxl tqdm
    Set ANTHROPIC_API_KEY environment variable.
"""

import argparse
import asyncio
import base64
import json
import os
import re
import sys
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from tqdm.asyncio import tqdm as atqdm

try:
    from anthropic import AsyncAnthropic
except ImportError:
    print("ERROR: anthropic package not installed. Run: pip install anthropic openpyxl tqdm")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
MODEL = "claude-opus-4-6"
MAX_FILE_SIZE_MB = 30

SUPPORTED_DOC_EXTENSIONS = {".pdf", ".jpg", ".jpeg", ".png", ".tiff", ".tif", ".bmp", ".webp"}

MANDATE_KEYWORDS = {"mandate", "debit", "dom", "ddo", "authoris", "authoriz"}
BANK_KEYWORDS = {"bank", "proof", "statement", "account", "confirmation", "letter", "confirm"}

# Words to strip when building a company keyword list for filename matching
COMPANY_STOPWORDS = {
    "pty", "ltd", "limited", "inc", "corp", "corporation", "the", "and", "of",
    "cc", "npc", "(pty)", "(ltd)", "proprietary",
}

MEDIA_TYPES = {
    ".pdf": "application/pdf",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".tiff": "image/tiff",
    ".tif": "image/tiff",
    ".bmp": "image/bmp",
    ".webp": "image/webp",
}

# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------

def find_header_row(ws) -> Optional[int]:
    """Find the row index (1-based) that contains 'Registered Company Name'."""
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "registered company name" in cell.value.lower():
                return cell.row
    return None


def read_companies_from_excel(excel_path: Path) -> list[dict]:
    """
    Parse the signup Excel file and return a list of company dicts:
      {fis_number, company_name, reg_number}
    Skips rows where company_name is blank.
    """
    companies = []
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as exc:
        return [{"fis_number": None, "company_name": None, "reg_number": None,
                 "error": f"Could not open Excel: {exc}"}]

    for sheet_name in wb.sheetnames:
        if "example" in sheet_name.lower():
            continue
        ws = wb[sheet_name]

        header_row = find_header_row(ws)
        if header_row is None:
            continue

        # Map header names to column indices (1-based)
        headers = {}
        for cell in ws[header_row]:
            if isinstance(cell.value, str):
                headers[cell.value.strip().lower()] = cell.column

        fis_col = headers.get("fis number")
        company_col = headers.get("registered company name")
        reg_col = headers.get("company registration number")

        if company_col is None:
            continue

        # Read data rows below header
        for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
            company_cell = row[company_col - 1]
            company_name = company_cell.value
            if not company_name or not str(company_name).strip():
                continue

            fis = row[fis_col - 1].value if fis_col else None
            reg = row[reg_col - 1].value if reg_col else None

            companies.append({
                "fis_number": str(fis).strip() if fis else None,
                "company_name": str(company_name).strip(),
                "reg_number": str(reg).strip() if reg else None,
            })

        if companies:
            break  # Use the first matching sheet only

    return companies


# ---------------------------------------------------------------------------
# File utilities
# ---------------------------------------------------------------------------

def get_media_type(path: Path) -> str:
    return MEDIA_TYPES.get(path.suffix.lower(), "application/octet-stream")


def encode_file(path: Path) -> str:
    with open(path, "rb") as f:
        return base64.standard_b64encode(f.read()).decode("utf-8")


def build_content_block(path: Path) -> dict:
    media_type = get_media_type(path)
    data = encode_file(path)
    if media_type == "application/pdf":
        return {
            "type": "document",
            "source": {"type": "base64", "media_type": "application/pdf", "data": data},
        }
    return {
        "type": "image",
        "source": {"type": "base64", "media_type": media_type, "data": data},
    }


def file_too_large(path: Path) -> bool:
    return path.stat().st_size > MAX_FILE_SIZE_MB * 1024 * 1024


def company_keywords(name: str) -> set[str]:
    """Extract meaningful lowercase keywords from a company name."""
    tokens = re.split(r"[\s\-_/\\()]+", name.lower())
    return {t for t in tokens if t and t not in COMPANY_STOPWORDS and len(t) > 1}


def filename_matches_company(filename: str, keywords: set[str]) -> bool:
    """Return True if any company keyword appears in the filename (case-insensitive)."""
    fname = filename.lower()
    return any(kw in fname for kw in keywords)


def classify_file_by_keyword(path: Path) -> str:
    """Return 'mandate', 'bank', or 'unknown' based on filename keywords."""
    stem = path.stem.lower()
    if any(kw in stem for kw in MANDATE_KEYWORDS):
        return "mandate"
    if any(kw in stem for kw in BANK_KEYWORDS):
        return "bank"
    return "unknown"


def get_doc_files(subfolder: Path) -> list[Path]:
    """Return all supported document files in the subfolder (no Excel files)."""
    return sorted(
        [
            f for f in subfolder.iterdir()
            if f.is_file()
            and f.suffix.lower() in SUPPORTED_DOC_EXTENSIONS
        ],
        key=lambda p: p.name.lower(),
    )


def assign_files_to_companies(
    companies: list[dict], doc_files: list[Path]
) -> list[tuple[dict, Optional[Path], Optional[Path]]]:
    """
    For each company, find its (mandate_file, bank_proof_file).

    Strategy:
    1. If only one company → split doc_files into mandates/banks by keyword,
       assign one of each. If only 1 file, classify it and mark the other missing.
    2. If multiple companies → match files to companies by company-name keywords
       in the filename; then within each match, split mandate vs bank.
    """
    if not companies:
        return []

    if len(companies) == 1:
        company = companies[0]
        mandates = [f for f in doc_files if classify_file_by_keyword(f) == "mandate"]
        banks = [f for f in doc_files if classify_file_by_keyword(f) == "bank"]
        unknown = [f for f in doc_files if classify_file_by_keyword(f) == "unknown"]

        # Promote unknowns: if only one category is empty, the unknowns fill it
        if not mandates and unknown:
            mandates, unknown = unknown[:1], unknown[1:]
        if not banks and unknown:
            banks, unknown = unknown[:1], unknown[1:]

        return [(company, mandates[0] if mandates else None, banks[0] if banks else None)]

    # Multiple companies — match by keyword then classify within matches
    result = []
    unassigned = list(doc_files)

    for company in companies:
        kws = company_keywords(company["company_name"])
        matched = [f for f in unassigned if filename_matches_company(f.name, kws)]
        for m in matched:
            unassigned.remove(m)

        mandate = next((f for f in matched if classify_file_by_keyword(f) == "mandate"), None)
        bank = next((f for f in matched if classify_file_by_keyword(f) == "bank"), None)

        # Promote unclassified matches
        for f in matched:
            if classify_file_by_keyword(f) == "unknown":
                if mandate is None:
                    mandate = f
                elif bank is None:
                    bank = f

        result.append((company, mandate, bank))

    # Distribute any remaining unassigned files to companies still missing docs
    for company, mandate, bank in result:
        pass  # covered below by re-building result

    # Rebuild with unassigned distributed
    final = []
    unassigned_iter = iter(unassigned)
    for company, mandate, bank in result:
        if mandate is None:
            nxt = next(unassigned_iter, None)
            if nxt and classify_file_by_keyword(nxt) in ("mandate", "unknown"):
                mandate = nxt
        if bank is None:
            nxt = next(unassigned_iter, None)
            if nxt and classify_file_by_keyword(nxt) in ("bank", "unknown"):
                bank = nxt
        final.append((company, mandate, bank))
    return final


# ---------------------------------------------------------------------------
# Claude API calls
# ---------------------------------------------------------------------------

def _parse_json_response(text: str) -> dict:
    text = re.sub(r"```(?:json)?\s*", "", text).strip().rstrip("`").strip()
    # Find the first { ... } block
    match = re.search(r"\{.*\}", text, re.DOTALL)
    if match:
        text = match.group(0)
    return json.loads(text)


async def analyze_mandate(
    client: AsyncAnthropic, path: Path, company_name: str
) -> dict:
    """
    Ask Claude:
    - Is the mandate signed?
    - What bank account number is on it?
    - (Optionally) does it relate to the given company?
    """
    try:
        response = await client.messages.create(
            model=MODEL,
            max_tokens=512,
            messages=[{
                "role": "user",
                "content": [
                    build_content_block(path),
                    {
                        "type": "text",
                        "text": (
                            f"This should be a debit order mandate for the company: {company_name}.\n\n"
                            "Analyze the document and return ONLY valid JSON (no markdown) with these fields:\n"
                            "{\n"
                            '  "is_signed": true or false,\n'
                            '  "bank_account_number": "account number as string or null",\n'
                            '  "signature_details": "one sentence on signature status"\n'
                            "}\n\n"
                            "Set is_signed to true only if a visible handwritten or electronic signature "
                            "is present in the designated signature field. "
                            "An empty box, typed name, or printed text does not count as a signature."
                        ),
                    },
                ],
            }],
        )
        text = next((b.text for b in response.content if b.type == "text"), "{}")
        return _parse_json_response(text)
    except json.JSONDecodeError as exc:
        return {"is_signed": False, "bank_account_number": None,
                "signature_details": f"JSON parse error: {exc}"}
    except Exception as exc:
        return {"is_signed": False, "bank_account_number": None,
                "signature_details": f"API error: {exc}"}


async def analyze_bank_proof(
    client: AsyncAnthropic, path: Path, company_name: str, mandate_account: Optional[str]
) -> dict:
    """
    Ask Claude:
    - What date is on the document?
    - Is it January 2025 or more recent?
    - Does the account number match the mandate's account number?
    """
    account_context = (
        f"The bank account number from the debit order mandate is: {mandate_account}"
        if mandate_account
        else "No account number was found on the debit order mandate."
    )
    try:
        response = await client.messages.create(
            model=MODEL,
            max_tokens=512,
            messages=[{
                "role": "user",
                "content": [
                    build_content_block(path),
                    {
                        "type": "text",
                        "text": (
                            f"This should be a proof of bank account for: {company_name}.\n"
                            f"{account_context}\n\n"
                            "Analyze the document and return ONLY valid JSON (no markdown) with these fields:\n"
                            "{\n"
                            '  "document_date": "YYYY-MM-DD or null if not found",\n'
                            '  "is_january_2025_or_later": true or false,\n'
                            '  "account_number_on_document": "account number or null",\n'
                            '  "account_numbers_match": true, false, or null,\n'
                            '  "date_details": "one sentence describing the date on the document"\n'
                            "}\n\n"
                            "Rules:\n"
                            "- is_january_2025_or_later: true if the document date is January 2025 or more recent; "
                            "false if older OR date cannot be determined.\n"
                            "- account_numbers_match: true if matching, false if different, null if either is unknown."
                        ),
                    },
                ],
            }],
        )
        text = next((b.text for b in response.content if b.type == "text"), "{}")
        return _parse_json_response(text)
    except json.JSONDecodeError as exc:
        return {"document_date": None, "is_january_2025_or_later": False,
                "account_number_on_document": None, "account_numbers_match": None,
                "date_details": f"JSON parse error: {exc}"}
    except Exception as exc:
        return {"document_date": None, "is_january_2025_or_later": False,
                "account_number_on_document": None, "account_numbers_match": None,
                "date_details": f"API error: {exc}"}


# ---------------------------------------------------------------------------
# Per-company validation
# ---------------------------------------------------------------------------

async def validate_company(
    client: AsyncAnthropic,
    subfolder_name: str,
    company: dict,
    mandate_file: Optional[Path],
    bank_file: Optional[Path],
) -> dict:
    """Validate one company's documents. Returns a result row dict."""
    result = {
        "subfolder": subfolder_name,
        "fis_number": company.get("fis_number") or "",
        "company_name": company.get("company_name") or "",
        "mandate_passed": False,
        "bank_passed": False,
        "failures": [],
    }

    company_name = company.get("company_name", "Unknown")
    mandate_account: Optional[str] = None

    # ── Mandate ────────────────────────────────────────────────────────────
    if mandate_file is None:
        result["failures"].append("Debit order mandate file not found in folder")
    elif file_too_large(mandate_file):
        result["failures"].append(f"Mandate file too large (>{MAX_FILE_SIZE_MB} MB): {mandate_file.name}")
    else:
        m = await analyze_mandate(client, mandate_file, company_name)
        mandate_account = m.get("bank_account_number")
        if m.get("is_signed"):
            result["mandate_passed"] = True
        else:
            result["failures"].append(f"Mandate not signed — {m.get('signature_details', '')}")

    # ── Bank proof ─────────────────────────────────────────────────────────
    if bank_file is None:
        result["failures"].append("Proof of bank account file not found in folder")
    elif file_too_large(bank_file):
        result["failures"].append(f"Bank proof file too large (>{MAX_FILE_SIZE_MB} MB): {bank_file.name}")
    else:
        b = await analyze_bank_proof(client, bank_file, company_name, mandate_account)
        bank_failures = []

        if not b.get("is_january_2025_or_later"):
            date_info = b.get("document_date") or "date not found"
            bank_failures.append(
                f"Bank proof older than January 2025 (document date: {date_info}; {b.get('date_details', '')})"
            )

        match = b.get("account_numbers_match")
        if match is False:
            bank_failures.append(
                f"Account number mismatch — mandate: {mandate_account or 'unknown'}, "
                f"bank proof: {b.get('account_number_on_document') or 'unknown'}"
            )
        elif match is None:
            bank_failures.append(
                "Could not verify account number match — one or both account numbers unreadable"
            )

        if bank_failures:
            result["failures"].extend(bank_failures)
        else:
            result["bank_passed"] = True

    return result


# ---------------------------------------------------------------------------
# Per-subfolder orchestration
# ---------------------------------------------------------------------------

async def process_subfolder(
    client: AsyncAnthropic, subfolder: Path, semaphore: asyncio.Semaphore
) -> list[dict]:
    """
    Process one subfolder:
    1. Find and parse the signup Excel.
    2. For each company, find mandate + bank files.
    3. Validate and return list of result rows.
    """
    async with semaphore:
        # Find Excel file (skip temp/lock files starting with ~$)
        excel_files = [
            f for f in subfolder.iterdir()
            if f.is_file()
            and f.suffix.lower() in {".xlsx", ".xls"}
            and not f.name.startswith("~$")
        ]

        if not excel_files:
            return [{
                "subfolder": subfolder.name,
                "fis_number": "",
                "company_name": "",
                "mandate_passed": False,
                "bank_passed": False,
                "failures": ["No signup Excel file found in folder"],
            }]

        companies = read_companies_from_excel(excel_files[0])

        if not companies:
            return [{
                "subfolder": subfolder.name,
                "fis_number": "",
                "company_name": "",
                "mandate_passed": False,
                "bank_passed": False,
                "failures": ["No company data found in Excel file (check 'Registered Company Name' column)"],
            }]

        # Check for Excel parse error
        if companies[0].get("error"):
            return [{
                "subfolder": subfolder.name,
                "fis_number": "",
                "company_name": "",
                "mandate_passed": False,
                "bank_passed": False,
                "failures": [companies[0]["error"]],
            }]

        doc_files = get_doc_files(subfolder)
        assignments = assign_files_to_companies(companies, doc_files)

        # Validate each company concurrently within the subfolder
        tasks = [
            validate_company(client, subfolder.name, company, mandate_f, bank_f)
            for company, mandate_f, bank_f in assignments
        ]
        return list(await asyncio.gather(*tasks))


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
PASS_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
PASS_FONT = Font(color="276221", bold=True)
FAIL_FONT = Font(color="9C0006", bold=True)
ALT_FILL = PatternFill(fill_type="solid", fgColor="F2F2F2")


def write_excel(all_results: list[dict], output_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Validation Results"

    headers = [
        "Subfolder", "FIS Number", "Registered Company Name",
        "Mandate Valid", "Bank Proof Valid", "Failure Description",
    ]
    col_widths = [20, 14, 35, 16, 16, 90]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[1].height = 22

    prev_subfolder = None
    alt = False

    for row_idx, r in enumerate(all_results, 2):
        # Alternate shading per subfolder group
        if r["subfolder"] != prev_subfolder:
            alt = not alt
            prev_subfolder = r["subfolder"]
        row_bg = ALT_FILL if alt else None

        ws.cell(row=row_idx, column=1, value=r["subfolder"]).alignment = Alignment(vertical="top")
        ws.cell(row=row_idx, column=2, value=r["fis_number"]).alignment = Alignment(vertical="top")
        ws.cell(row=row_idx, column=3, value=r["company_name"]).alignment = Alignment(vertical="top")

        if row_bg:
            for col in (1, 2, 3):
                ws.cell(row=row_idx, column=col).fill = row_bg

        for col, passed in [(4, r["mandate_passed"]), (5, r["bank_passed"])]:
            cell = ws.cell(row=row_idx, column=col, value="PASS" if passed else "FAIL")
            cell.fill = PASS_FILL if passed else FAIL_FILL
            cell.font = PASS_FONT if passed else FAIL_FONT
            cell.alignment = Alignment(horizontal="center", vertical="top")

        failure_text = "; ".join(r["failures"]) if r["failures"] else ""
        ws.cell(row=row_idx, column=6, value=failure_text).alignment = Alignment(
            wrap_text=True, vertical="top"
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{len(all_results) + 1}"
    wb.save(output_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

async def main(root_folder: str, output_file: str, concurrency: int) -> None:
    if not os.environ.get("ANTHROPIC_API_KEY"):
        print("ERROR: ANTHROPIC_API_KEY environment variable is not set.")
        sys.exit(1)

    root = Path(root_folder)
    if not root.exists() or not root.is_dir():
        print(f"ERROR: Folder does not exist: {root_folder}")
        sys.exit(1)

    subfolders = sorted([f for f in root.iterdir() if f.is_dir()], key=lambda p: p.name.lower())
    if not subfolders:
        print("No subfolders found.")
        return

    print(f"Found {len(subfolders)} subfolders.")
    print(f"Model  : {MODEL}")
    print(f"Output : {output_file}")
    print(f"Concurrency: {concurrency}\n")

    client = AsyncAnthropic()
    semaphore = asyncio.Semaphore(concurrency)

    tasks = [process_subfolder(client, sf, semaphore) for sf in subfolders]
    all_results: list[dict] = []

    async for coro in atqdm(asyncio.as_completed(tasks), total=len(tasks), desc="Processing folders"):
        rows = await coro
        all_results.extend(rows)

    # Sort: subfolder name first, then company name
    all_results.sort(key=lambda r: (r["subfolder"].lower(), r["company_name"].lower()))

    write_excel(all_results, output_file)

    total = len(all_results)
    m_pass = sum(1 for r in all_results if r["mandate_passed"])
    b_pass = sum(1 for r in all_results if r["bank_passed"])
    full = sum(1 for r in all_results if r["mandate_passed"] and r["bank_passed"])

    print("\n" + "-" * 52)
    print(f"  Companies processed     : {total}")
    print(f"  Mandate PASS            : {m_pass}/{total}")
    print(f"  Bank proof PASS         : {b_pass}/{total}")
    print(f"  Both passed             : {full}/{total}")
    print("-" * 52)
    print(f"  Results saved to: {output_file}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Validate debit order mandates and bank proofs.")
    parser.add_argument("--folder", "-f", required=True,
                        help="Root folder containing subfolders (each with an Excel + documents).")
    parser.add_argument("--output", "-o", default="validation_results.xlsx",
                        help="Output Excel file path.")
    parser.add_argument("--concurrency", "-c", type=int, default=5,
                        help="Concurrent folder processing (default: 5).")
    args = parser.parse_args()
    asyncio.run(main(args.folder, args.output, args.concurrency))

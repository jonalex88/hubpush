import os
from pathlib import Path
import openpyxl

root = Path(r'c:\Vibes\Hubpush\1.FBEO\Sample')
for folder in sorted([p for p in root.iterdir() if p.is_dir()]):
    print(f'\n== {folder.name} ==')
    xlsx = sorted(folder.glob('*.xlsx'))
    print('xlsx:', [x.name for x in xlsx])
    for xf in xlsx:
        wb = openpyxl.load_workbook(xf, data_only=True)
        for ws in wb.worksheets:
            # scan first 120 rows and all cols for likely header line
            max_r = min(ws.max_row, 120)
            found = False
            for r in range(1, max_r+1):
                vals = [ws.cell(r,c).value for c in range(1, ws.max_column+1)]
                row_text = ' | '.join('' if v is None else str(v) for v in vals)
                low = row_text.lower()
                if 'registered company name' in low or ('company' in low and 'name' in low):
                    print(f'  sheet={ws.title} header_row={r} text={row_text[:240]}')
                    found = True
                    break
            if not found:
                print(f'  sheet={ws.title} no obvious company header found')
        wb.close()

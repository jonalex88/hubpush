import openpyxl
p = r'c:\Vibes\Hubpush\output v2 all fields FULL.resume.xlsx'
wb = openpyxl.load_workbook(p, data_only=True)
ws = wb.active
print('ROWS', ws.max_row)
print('COLS', ws.max_column)
print('HEADERS', [ws.cell(1,c).value for c in range(1, min(ws.max_column,12)+1)])
wb.close()

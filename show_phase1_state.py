import json
import openpyxl
from collections import Counter

master = r'c:\Vibes\Hubpush\output v2 all fields FULL.resume.xlsx'
snap = r'c:\Vibes\Hubpush\data\cloud_master_snapshot.json'
commits = r'c:\Vibes\Hubpush\data\commit_history.json'

wb = openpyxl.load_workbook(master, data_only=True)
ws = wb.active
headers = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
idx = {h:i+1 for i,h in enumerate(headers)}
status_col = idx.get('HubSpot Status')
rowid_col = idx.get('Cloud Row ID')
checksum_col = idx.get('Row Checksum')

statuses = []
row_ids = []
for r in range(2, ws.max_row+1):
    company = ws.cell(r, idx.get('company reg name',2)).value
    if not company:
        continue
    statuses.append(str(ws.cell(r, status_col).value or '').strip())
    row_ids.append(str(ws.cell(r, rowid_col).value or '').strip())

wb.close()

with open(snap, 'r', encoding='utf-8') as f:
    snap_obj = json.load(f)
with open(commits, 'r', encoding='utf-8') as f:
    commits_obj = json.load(f)

print('MASTER_FILE', master)
print('ROWS_IN_MASTER', len(statuses))
print('UNIQUE_ROW_IDS', len(set(row_ids)))
print('HAS_CHECKSUM_COLUMN', checksum_col is not None)
print('STATUS_COUNTS', dict(Counter(statuses)))
print('SNAPSHOT_ROW_COUNT', snap_obj.get('row_count'))
print('SNAPSHOT_UPDATED_AT', snap_obj.get('updated_at'))
print('COMMIT_HISTORY_COUNT', len(commits_obj.get('commits', [])))

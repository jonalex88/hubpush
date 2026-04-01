import openpyxl
wb = openpyxl.load_workbook(r'c:\Vibes\Hubpush\output v2 all fields FULL.resume.xlsx', data_only=True)
ws = wb.active
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
print("HEADERS:", headers)
# Find first row with both PASS
for r in range(2, ws.max_row + 1):
    mval = ws.cell(r, 5).value
    bval = ws.cell(r, 6).value
    if mval == 'PASS' and bval == 'PASS':
        row = {headers[c-1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        print("FIRST_PASS_ROW:", row)
        break
wb.close()

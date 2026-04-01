import openpyxl
from pathlib import Path
root = Path(r'c:\Vibes\Hubpush\1.FBEO\Sample')
for x in root.rglob('*.xlsx'):
    print('\nFILE', x.name)
    wb=openpyxl.load_workbook(x,data_only=True)
    for ws in wb.worksheets:
        max_r=min(ws.max_row,40)
        max_c=min(ws.max_column,40)
        for r in range(1,max_r+1):
            vals=[ws.cell(r,c).value for c in range(1,max_c+1)]
            txt=' | '.join('' if v is None else str(v) for v in vals)
            low=txt.lower()
            if 'registered company name' in low or 'company registration' in low:
                print(' SHEET',ws.title,'ROW',r)
                print(' ',txt)
                # show next 8 rows for company columns
                for rr in range(r+1,min(r+9,max_r+1)):
                    vals2=[ws.cell(rr,c).value for c in range(1,max_c+1)]
                    print('   ',rr,vals2)
                break
    wb.close()
